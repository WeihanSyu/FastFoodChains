### Open our raw text files containing the menu info and extract what we need from them ######################################################
# Since we are only using skipthedishes.com, we expect every restaurant page layout to be the same, but they have small inconsistencies here and there
import re
import pandas as pd
from openpyxl import load_workbook


### Function to clean and organize any menu taken from Skipthedishes ####################################################################################################################
def clean_menu_skipdishes(filepath, encode="utf-8"):
    f = open(filepath, encoding=encode)
    menu = []
    for line in f:
        menu.append(line)
    f.close()

    # Remove all leading/trailing blank spaces & newline elements from the menu list
    for i, x in enumerate(menu):
        menu[i].strip()
        menu[i] = re.sub(r'\n', '', x)

    # Reconstruct the menu list and avoid any empty strings - which give us False when checked in (if value) statement
    menu = [x for x in menu if x]
    
    # Remove the duplicate item names due to getting the plain text of images having the same name
    price_indicies = [i for i, x in enumerate(menu) if re.search('\$[0-9]+\.[0-9]+', x)]
    j = 0
    for i in price_indicies:
        if menu[i - 2 - j] == menu[i - 3 - j]:
            menu.pop(i - 2 - j)
            j += 1

    see_item_indicies = [i for i, x in enumerate(menu) if re.search('See Item', x)]
    j = 0
    for i in see_item_indicies:
        if menu[i - 2 - j] == menu[i - 3 - j]:
            menu.pop(i - 2 - j)
            j += 1

    sold_indicies = [i for i, x in enumerate(menu) if re.search('SOLD OUT', x)]
    j = 0
    for i in sold_indicies:
        if menu[i - 2 - j] == menu[i - 3 - j]:
            menu.pop(i - 2 - j)
            j += 1
    
    # Remove the item description right before every price element. Note that some franchises place the calories in this section. So we need it.
    price_indicies = [i for i, x in enumerate(menu) if re.search('\$[0-9]+\.[0-9]+', x)]
    j = 0
    for i in price_indicies:
        # KFC format
        if re.search('\([0-9]+-*[0-9]* Cals/Person\)', menu[i - 1 - j]):
            cals = re.findall('\([0-9]+-*[0-9]* Cals/Person\)', menu[i - 1 - j])[0]
            menu[i - 2 - j] += cals   # Append calories to food item name which is always before the description
        # Subway format
        if re.search('\[[0-9]+ Cals\]', menu[i - 1 - j]):
            cals = re.findall('\[[0-9]+ Cals\]', menu[i - 1 - j])[0]
            menu[i - 2 - j] += cals

        menu.pop(i - 1 - j)
        j += 1
        # Since our list gets shorter each time we remove an element, we just decrease our price indicies by 1 every time as well

    # Do the same with 'See Item'. 
    see_item_indicies = [i for i, x in enumerate(menu) if re.search('See Item', x)]
    j = 0
    for i in see_item_indicies:
        # KFC format
        if re.search('\([0-9]+-*[0-9]* Cals/Person\)', menu[i - 1 - j]):
            cals = re.findall('\([0-9]+-*[0-9]* Cals/Person\)', menu[i - 1 - j])[0]
            menu[i - 2 - j] += cals   # Append calories to food item name which is always before the description
        # Subway format
        if re.search('\[[0-9]+ Cals\]', menu[i - 1 - j]):
            cals = re.findall('\[[0-9]+ Cals\]', menu[i - 1 - j])[0]
            menu[i - 2 - j] += cals

        menu.pop(i - 1 - j)
        j += 1

    # And with SOLD OUT
    sold_indicies = [i for i, x in enumerate(menu) if re.search('SOLD OUT', x)]
    j = 0
    for i in sold_indicies:
        # KFC format
        if re.search('\([0-9]+-*[0-9]* Cals/Person\)', menu[i - 1 - j]):
            cals = re.findall('\([0-9]+-*[0-9]* Cals/Person\)', menu[i - 1 - j])[0]
            menu[i - 2 - j] += cals   # Append calories to food item name which is always before the description
        # Subway format
        if re.search('\[[0-9]+ Cals\]', menu[i - 1 - j]):
            cals = re.findall('\[[0-9]+ Cals\]', menu[i - 1 - j])[0]
            menu[i - 2 - j] += cals

        menu.pop(i - 1 - j)
        j += 1
    
    # Remove the Section Titles from the menu. Some have two lines of consequtive titles and some only have one line. So check it twice.
    price_indicies = [i for i, x in enumerate(menu) if re.search('\$[0-9]+\.[0-9]+', x)]
    j = 0
    for i in price_indicies:
        if not re.search('\$[0-9]+\.[0-9]+', menu[i - 2 - j]) and not re.search('See Item', menu[i - 2 - j]) and not re.search('SOLD OUT', menu[i - 2 - j]):
            menu.pop(i - 2 - j)
            j += 1
        if not re.search('\$[0-9]+\.[0-9]+', menu[i - 2 - j]) and not re.search('See Item', menu[i - 2 - j]) and not re.search('SOLD OUT', menu[i - 2 - j]):
            menu.pop(i - 2 - j)
            j += 1
        
    # Technically "See Item" takes the place of price so we have to check for it as well.
    see_item_indicies = [i for i, x in enumerate(menu) if re.search('See Item', x)]
    j = 0
    for i in see_item_indicies:
        if not re.search('\$[0-9]+\.[0-9]+', menu[i - 2 - j]) and not re.search('See Item', menu[i - 2 - j]) and not re.search('SOLD OUT', menu[i - 2 - j]):
            menu.pop(i - 2 - j)
            j += 1
        if not re.search('\$[0-9]+\.[0-9]+', menu[i - 2 - j]) and not re.search('See Item', menu[i - 2 - j]) and not re.search('SOLD OUT', menu[i - 2 - j]):
            menu.pop(i - 2 - j)
            j += 1

    # Subway and Mcdonalds also includes SOLD OUT items in place of prices. 
    sold_indicies = [i for i, x in enumerate(menu) if re.search('SOLD OUT', x)]
    j = 0
    for i in sold_indicies:
        if not re.search('\$[0-9]+\.[0-9]+', menu[i - 2 - j]) and not re.search('See Item', menu[i - 2 - j]) and not re.search('SOLD OUT', menu[i - 2 - j]):
            menu.pop(i - 2 - j)
            j += 1
        if not re.search('\$[0-9]+\.[0-9]+', menu[i - 2 - j]) and not re.search('See Item', menu[i - 2 - j]) and not re.search('SOLD OUT', menu[i - 2 - j]):
            menu.pop(i - 2 - j)
            j += 1

    return menu


### Function to clean and organize any menu taken from UberEats ##########################################################################################################################
def clean_menu_ubereats(filepath, encode="utf-8"):
    f = open(filepath, encoding=encode)
    menu = []
    for line in f:
        menu.append(line)
    f.close()

    ### NEW APPROACH TO AVOID USING .pop(). Just turn everything you don't want into an empty string. Then reconstruct the list without any empty strings #############

    # Remove all leading/trailing blank spaces & newline elements from the menu list
    for i, x in enumerate(menu):
        menu[i].strip()
        menu[i] = re.sub(r'\n', '', x)

    # Remove weird ' • ' elements. 
    for i, x in enumerate(menu):
        if x == ' • ':
            menu[i] = ''

    # Remove any user ratings
    for i, x in enumerate(menu):
        if re.search("[0-9]+\% \([0-9]+\)", x):
            menu[i] = ''

    # Reconstruct the menu list and avoid any empty strings - which give us False when checked in (if value) statement
    menu = [x for x in menu if x]

    # Check if the calorie content line exists (some restaurants do not include), then remove item descriptions + duplicate item names accordingly
    price_indicies = [i for i, x in enumerate(menu) if re.search('US\$[0-9]+\.[0-9]+', x)]
    for i in price_indicies:
        if re.search("[0-9]+ Cal", menu[i + 1]):   
            if menu[i - 1] != menu[i + 2]:   # Sometimes there is a description after the calories
                menu[i + 2] = ''
                menu[i + 3] = ''
            else:                           
                menu[i + 2] = ''
        else:
            if menu[i - 1] != menu[i + 1]:  # Sometimes there is a description after price (ratings removed) 
                menu[i + 1] = ''
                menu[i + 2] = ''
            else:
                menu[i + 1] = ''

    # Remove empty strings again
    menu = [x for x in menu if x]

    # Remove section titles
    price_indicies = [i for i, x in enumerate(menu) if re.search('US\$[0-9]+\.[0-9]+', x)]
    for i in price_indicies:
        if not re.search("[0-9]+ Cal", menu[i - 2]) and not re.search('US\$[0-9]+\.[0-9]+', menu[i - 2]):
            menu[i - 2] = ''
    menu = [x for x in menu if x]

    return menu


### Function to clean and organize any calorie info from the official websites ###########################################################################################################
def clean_cals(filepath, encode="utf-8"):
    f = open(filepath, encoding=encode)
    menu = []
    for line in f:
        menu.append(line)
    f.close()

     # Remove all leading/trailing blank spaces & newline elements from the menu list
    for i, x in enumerate(menu):
        menu[i].strip()
        menu[i] = re.sub(r'\n', '', x)

    # Reconstruct the menu list and avoid any empty strings - which give us False when checked in (if value) statement
    menu = [x for x in menu if x]

    return menu


### Function to take any menu from Skipthedishes OR UberEats and load it to Excel ########################################################################################################
def menu_to_excel(menu, filename):
    if re.search("[0-9]+ Cal", menu[2]):    # If there are three repeating groups (name, price, calories)
        # Use DOUBLE COLON index slicing [start:stop:steps] to get every nth element in the list
        item_name = menu[::3]
        price = menu[1::3]
        calories = menu[2::3]
        price = [float(re.sub("US\$", "", x)) if re.search('\$[0-9]+\.[0-9]+', x) else x for x in price]     # Get rid of non-numeric characters for the price and convert to float
        price = [round(x / 1.2, 2) if type(x) is float else x for x in price]    # Apply price reduction here to account for Skip/Uber fees
        calories = [re.sub(" Cal[.]*", "", x) for x in calories]
        
        df = pd.DataFrame({
            'item_name': item_name,
            'price': price,
            'calories': calories
        })
        
    else:   # only name and price (skip the dishes calories always appended to name)
        item_name = [x for i, x in enumerate(menu) if i % 2 == 0]
        price = [x for i, x in enumerate(menu) if i % 2 != 0]   
        price = [float(re.sub("US\$", "", x)) if re.search('\$[0-9]+\.[0-9]+', x) else x for x in price]     # Get rid of the dollar sign and convert to float
        price = [round(x / 1.2, 2) if type(x) is float else x for x in price]    # Apply price reduction here to account for Skip/Uber fees
        
        df = pd.DataFrame({
            'item_name': item_name,
            'price': price
        })

    try:
        existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Clean\\menus\\' + filename + '.xlsx')
        writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Clean\\menus\\' + filename + '.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
        startrow = writer.sheets['Sheet1'].max_row
        df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
        writer.close()
    except Exception as ex:
        print(ex)
        df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Clean\\menus\\' + filename + '.xlsx', index=False)


### Function to take any calorie list and load it to Excel ################################################################################################################################
def cals_to_excel(menu, filename):
    item_name = [x for i, x in enumerate(menu) if i % 2 == 0]
    cals = [x for i, x in enumerate(menu) if i % 2 != 0]   
    cals = [re.sub("[a-zA-Z]+", "", x).strip() for x in cals]
    
    df = pd.DataFrame({
        'item_name': item_name,
        'calories': cals
    })

    try:
        existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Clean\\menus\\calories\\' + filename + '.xlsx')
        writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Clean\\menus\\calories\\' + filename + '.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
        startrow = writer.sheets['Sheet1'].max_row
        df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
        writer.close()
    except Exception as ex:
        print(ex)
        df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Clean\\menus\\calories\\' + filename + '.xlsx', index=False)






# For KFC we need to remove the duplicate item names first, otherwise appending calories to one of them will cause them to no longer be duplicates

# For Subway, I think it worked because our titles were only 1 line each while KFC was 2. and we are taking all odd and all even numbers so it worked out
# We should still redo it though.


### END RESULT: Two columns, 'item_name', 'price'. Calories are appended to item_name. Duplicates have not been sorted out. ###################################################################
