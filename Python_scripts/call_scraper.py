### This file is for calling the scrapers and returning their data into Excel files #####################################################################################################

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import pandas as pd
from openpyxl import load_workbook

from construct_scraper import scrape_construct


service = Service(executable_path='C:\\ChromeDrive\\chromedriver.exe')
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_argument("--incognito")
chrome_options.add_experimental_option("detach", True)

# For brave browser instead of chrome
brave_path = r"C:/Program Files/BraveSoftware/Brave-Browser/Application/brave.exe"


class call_scrape:
    def call_subway_CA(self):
        run = scrape_construct(service, chrome_options)
        raw_text, province, city, address, postcode = run.subway_CA()
        
        df = pd.DataFrame({
            'raw_text': raw_text,
            'state_province': province,
            'city': city,
            'address': address,
            'postcode': postcode
        })

        try:
            existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\subwayCA.xlsx')
            writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\subwayCA.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
            startrow = writer.sheets['Sheet1'].max_row
            df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
            writer.close()
        except Exception as ex:
            print(ex)
            df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\subwayCA.xlsx', index=False)


    def call_kfc_CA(self):
        run = scrape_construct(service, chrome_options)
        raw_text = run.kfc_CA()

        df = pd.DataFrame({'full_address': raw_text})

        try:
            existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\kfcCA.xlsx')
            writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\kfcCA.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
            startrow = writer.sheets['Sheet1'].max_row
            df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
            writer.close()
        except Exception as ex:
            print(ex)
            df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\kfcCA.xlsx', index=False)


    def call_kfc_US(self):
        # Set web browser to use brave instead of chrome
        chrome_options.binary_location = brave_path

        run = scrape_construct(service, chrome_options)
        raw_text, state, city, address, postcode = run.kfc_US()

        df = pd.DataFrame({
            'raw_text': raw_text,
            'state_province': state,
            'city': city,
            'address': address,
            'postcode': postcode
        })

        try:
            existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\kfcUS.xlsx')
            writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\kfcUS.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
            startrow = writer.sheets['Sheet1'].max_row
            df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
            writer.close()
        except Exception as ex:
            print(ex)
            df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\kfcUS.xlsx', index=False)


    def call_timhortons_CA(self):
        # Set web browser to use brave instead of chrome
        chrome_options.binary_location = brave_path

        run = scrape_construct(service, chrome_options)
        raw_text, province, city, address, postcode = run.timhortons_CA()

        df = pd.DataFrame({
            'raw_text': raw_text,
            'state_province': province,
            'city': city,
            'address': address,
            'postcode': postcode
        })

        try:
            existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\timhortonsCA.xlsx')
            writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\timhortonsCA.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
            startrow = writer.sheets['Sheet1'].max_row
            df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
            writer.close()
        except Exception as ex:
            print(ex)
            df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\timhortonsCA.xlsx', index=False)


    def call_timhortons_US(self):
        # Set web browser to use brave instead of chrome
        chrome_options.binary_location = brave_path

        run = scrape_construct(service, chrome_options)
        raw_text, state, city, address, postcode = run.timhortons_US()

        df = pd.DataFrame({
            'raw_text': raw_text,
            'state_province': state,
            'city': city,
            'address': address,
            'postcode': postcode
        })

        try:
            existing_data = load_workbook('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\timhortonsUS.xlsx')
            writer = pd.ExcelWriter('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\timhortonsUS.xlsx', mode='a', engine="openpyxl", if_sheet_exists='overlay')
            startrow = writer.sheets['Sheet1'].max_row
            df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, header=None, index=False)
            writer.close()
        except Exception as ex:
            print(ex)
            df.to_excel('C:\\Users\\Weihan\\PersonalProjects\\FastFoodChains\\Datasets\\Raw\\locations\\timhortonsUS.xlsx', index=False)


        

# test = call_scrape()
# test.call_timhortons_US()